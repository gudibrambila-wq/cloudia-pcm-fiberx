"""
gerar_dados.py — Extrai PSI_integrado_v19.xlsx e gera dados.json para o CloudIA
Execute: python gerar_dados.py
"""
import json
import os
import datetime
import openpyxl

# ---------------------------------------------------------------------------
# Configuração
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_DIR = os.path.join(BASE_DIR, "..")  # pasta Cloud PCM
EXCEL_FILE = os.path.join(EXCEL_DIR, "PSI_integrado_v19.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "dados.json")
HISTORICO_FILE = os.path.join(BASE_DIR, "historico_saldo.json")  # saldo congelado de semanas fechadas

# Base do calendário ISO (segunda-feira da W1)
BASE_DATE = datetime.date(2025, 12, 29)
# Calcula SEMANA_ATUAL automaticamente com base na data de hoje.
# Se precisar travar em uma semana específica, edite manualmente (comente a linha abaixo).
SEMANA_ATUAL = max(1, (datetime.date.today() - BASE_DATE).days // 7 + 1)
# SEMANA_ESTOQUE = semana da foto do Sênior. Por padrão = SEMANA_ATUAL, mas pode
# ser sobrescrito se a foto for de semana anterior (ex: baixado no fim de semana).
SEMANA_ESTOQUE = SEMANA_ATUAL

# Códigos cujo saldo NÃO deve usar a foto do V11 consolidado (têm outra fonte de verdade).
# Quando vazio, todos os 65 produtos usam a foto real consolidada (OMIE + Sênior + ML).
EXCECOES_FOTO_SENIOR = set()  # vazio: Robô e Alimentador também usam V11 agora

# Realizados manuais — sobrescreve psi.realizado[W] pra produtos específicos.
# Útil quando a semana já fechou mas o Excel-fonte ainda não tem o número atualizado.
# Formato: { codigo_senior: { semana: qtd } }
REALIZADO_MANUAL = {
    # Semanas fechadas com vendas reais (ML/outras fontes não B2B). Passar Douglas
    # semanalmente até termos base automatizada. Formato: { codigo: { semana: qtd_total } }
    400600001: { 20: 165, 21: 137, 22: 135, 23: 176, 24: 127, 25: 116, 26: 108, 27: 152, 28: 193, 29: 148 },  # Alimentador APBW
    400700001: { 20: 148, 21: 142, 22: 121, 23: 162, 24: 143, 25: 141, 26: 107, 27: 165, 28: 291, 29: 141 },  # Robô RALWC
    400700012: { 26:   3, 27:  29, 28:  71, 29:  64 },  # Robô V50
    400700013: { 26:   8, 27:  18, 28:  18, 29:   5 },  # Robô V50 PRO 110V
    400700014: { 26:   6, 27:  17, 28:   3, 29:   2 },  # Robô V50 PRO 220V
}

# Split do realizado por (semana, mês) — só pra semanas-fronteira (ex: W27 = JUN+JUL).
# Formato: { codigo: { "<semana>_<mes>": qtd } }
# Usado pra popular realizado_split (que o cliente exibe na linha "Realizado" da tabela,
# separando a semana-fronteira entre as 2 colunas de mês). Total tem que bater com
# REALIZADO_MANUAL do mesmo (codigo, semana).
REALIZADO_SPLIT_MANUAL = {
    400600001: { '27_JUN/26': 30,  '27_JUL/26': 122 },  # Alimentador — total 152 ✓
    400700001: { '27_JUN/26': 24,  '27_JUL/26': 141 },  # RALWC       — total 165 ✓
    400700012: { '27_JUN/26':  6,  '27_JUL/26':  23 },  # V50         — total  29 ✓
    400700013: { '27_JUN/26':  5,  '27_JUL/26':  13 },  # V50 PRO 110 — total  18 ✓
    400700014: { '27_JUN/26':  4,  '27_JUL/26':  13 },  # V50 PRO 220 — total  17 ✓
}

# Merge de cadastros que representam o MESMO produto com MESMO custo (só nome
# de cadastro diferente por questão de importadora/empresa). Ter "IMP" no nome
# NÃO é gatilho — TP-Link IMP tem custo diferente e continua produto separado.
# Só entra aqui quando Douglas confirma "é o mesmo item".
# Formato: { codigo_secundario: codigo_principal }
CODIGOS_MESCLADOS = {
    400600002: 400600001,  # Alimentador APBW — só diferença de nome de cadastro
    400700009: 400700001,  # Robô RALWC     — só diferença de nome de cadastro
}

# Mapeamento semana → mês (igual ao WEEK_TO_MONTH do index.html)
WEEK_TO_MONTH = {
    1: "JAN/26",  5: "FEV/26",  9: "MAR/26", 14: "ABR/26", 18: "MAI/26",
   23: "JUN/26", 27: "JUL/26", 31: "AGO/26", 36: "SET/26", 40: "OUT/26",
   44: "NOV/26", 49: "DEZ/26", 53: "JAN/27", 58: "FEV/27", 62: "MAR/27",
   66: "ABR/27", 70: "MAI/27", 75: "JUN/27", 79: "JUL/27", 83: "AGO/27",
   88: "SET/27", 92: "OUT/27", 97: "NOV/27", 101: "DEZ/27",
}

def semana_de_data(data: datetime.date) -> int:
    """Converte uma data em número de semana PSI."""
    if isinstance(data, datetime.datetime):
        data = data.date()
    return max(1, (data - BASE_DATE).days // 7 + 1)

def data_de_semana(semana: int) -> datetime.date:
    return BASE_DATE + datetime.timedelta(weeks=semana - 1)

def semanas_do_mes(mes: str) -> int:
    """Número real de semanas PSI do mês (considera meses com 4 ou 5 semanas)."""
    mm = {'JAN':1,'FEV':2,'MAR':3,'ABR':4,'MAI':5,'JUN':6,
          'JUL':7,'AGO':8,'SET':9,'OUT':10,'NOV':11,'DEZ':12}
    m_abr, y_suf = mes.split('/')
    year, month = 2000 + int(y_suf), mm[m_abr]
    if month == 12:
        ultimo_dia = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
    else:
        ultimo_dia = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)
    semana_fim = semana_de_data(ultimo_dia)
    semana_ini = next((w for w, m in sorted(WEEK_TO_MONTH.items()) if m == mes), None)
    return (semana_fim - semana_ini + 1) if semana_ini else 4

def parse_week_label(label) -> int | None:
    """Converte 'W14' → 14, ou retorna None."""
    if not isinstance(label, str):
        return None
    label = label.strip()
    if label.startswith("W") or label.startswith("w"):
        try:
            return int(label[1:])
        except ValueError:
            return None
    return None

# ---------------------------------------------------------------------------
# Carregar workbook
# ---------------------------------------------------------------------------
print(f"Lendo: {EXCEL_FILE}")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

# Carrega histórico de saldo (semanas fechadas congeladas em runs anteriores)
HISTORICO_SALDO = {}
if os.path.exists(HISTORICO_FILE):
    try:
        with open(HISTORICO_FILE, "r", encoding="utf-8") as f:
            HISTORICO_SALDO = json.load(f)
        print(f"  Histórico de saldo carregado: {len(HISTORICO_SALDO)} produtos")
    except Exception as e:
        print(f"  AVISO: não consegui ler {HISTORICO_FILE}: {e}")

# ---------------------------------------------------------------------------
# 1. Lista — produtos, preços, offsets e forecast mensal
# ---------------------------------------------------------------------------
ws_lista = wb["Lista"]
lista_headers = [c.value for c in next(ws_lista.iter_rows(max_row=1))]
# Índices das colunas mensais de forecast e venda
fcst_cols = {h: i for i, h in enumerate(lista_headers) if isinstance(h, str) and h.startswith("_fcst_")}
venda_cols = {h: i for i, h in enumerate(lista_headers) if isinstance(h, str) and h.startswith("_venda_")}

produtos_raw = []
for row in ws_lista.iter_rows(min_row=2, values_only=True):
    nome = row[0]
    if not nome:
        continue
    fcst_mensal = {}
    for col_key, col_idx in fcst_cols.items():
        mes_key = col_key.replace("_fcst_", "")  # ex: "JAN/26"
        val = row[col_idx]
        fcst_mensal[mes_key] = int(val) if isinstance(val, (int, float)) and val else 0

    venda_mensal = {}
    for col_key, col_idx in venda_cols.items():
        mes_key = col_key.replace("_venda_", "")
        val = row[col_idx]
        venda_mensal[mes_key] = int(val) if isinstance(val, (int, float)) and val else 0

    produtos_raw.append({
        "nome":          row[0],
        "marca":         row[1],
        "codigo":        row[2],
        "familia":       row[3],
        "_linha_psi":    row[4],
        "_off_fcst":     row[12] if row[12] else 3,
        "_off_sin":      row[13] if row[13] else 4,
        "_off_real":     row[14] if row[14] else 5,
        "_off_sld":      row[16] if row[16] else 7,
        "pv":            float(row[6]) if isinstance(row[6], (int, float)) else 0.0,
        "pc":            float(row[7]) if isinstance(row[7], (int, float)) else 0.0,
        "fcst_mensal":   fcst_mensal,
        "venda_mensal":  venda_mensal,
    })

print(f"  {len(produtos_raw)} produtos na Lista")

# ---------------------------------------------------------------------------
# 2. Tipo (Importado / Nacional) — via Planejamento_Compra
# ---------------------------------------------------------------------------
ws_plan = wb["Planejamento_Compra"]
tipo_por_produto = {}
for row in ws_plan.iter_rows(min_row=3, values_only=True):
    nome, marca, tipo = row[0], row[1], row[2]
    if nome and tipo and nome not in tipo_por_produto:
        tipo_por_produto[nome] = tipo  # "Importado" ou "Nacional"

# ---------------------------------------------------------------------------
# 3. ESTOQUE CONSOLIDADO V11 — fonte única (Velds CONSOLIDADO + Estoque FX - TELECOM)
# Gerado por consolidar_estoque_v19.py. Já tem RMA, peças e amostras filtradas.
# Se o V11 não existir, faz fallback para ler ESTOQUE SENIOR + ESTOQUE OMIE direto.
# ---------------------------------------------------------------------------
# Fonte de estoque: prioridade
#   1º) posicao_estoque_todas_<data>.xlsx — formato novo (3 empresas em 1 arquivo)
#       Gerado pelo Sênior direto, contém: Produto, Descrição, Família, Marca, Empresa,
#       Filial, Depósito, UN, Físico, Disponível, Saldo Final, Preço Médio, etc.
#   2º) Estoque FIBERX <data>.xlsx — formato antigo (V11/V12 consolidado)
#   3º) estoque_consolidada V11.xlsx — fallback root
_PASTA_POSICAO = os.path.join(BASE_DIR, "..", "Posição estoques")
POSICAO_TODAS_PATH = None
V12_CANDIDATOS = []
if os.path.isdir(_PASTA_POSICAO):
    _todos_arquivos = os.listdir(_PASTA_POSICAO)
    # Formato NOVO: posicao_estoque_todas_YYYY-MM-DD.xlsx (ignora _COM_RESUMO)
    _pos_todas = [os.path.join(_PASTA_POSICAO, f) for f in _todos_arquivos
                  if f.startswith("posicao_estoque_todas_") and f.endswith(".xlsx")
                  and "COM_RESUMO" not in f.upper()]
    # Pega o mais recente que CONSEGUE abrir (pula se estiver bloqueado no Excel)
    _pos_todas.sort(key=os.path.getmtime, reverse=True)
    POSICAO_TODAS_PATH = None
    for _p in _pos_todas:
        try:
            with open(_p, "rb"): pass
            POSICAO_TODAS_PATH = _p
            break
        except PermissionError:
            print(f"  [AVISO] pulando {os.path.basename(_p)} (arquivo bloqueado)")
    # Formato ANTIGO como fallback
    V12_CANDIDATOS = [os.path.join(_PASTA_POSICAO, f) for f in _todos_arquivos
                      if f.lower().startswith("estoque fiberx") and f.endswith(".xlsx")]
    V12_CANDIDATOS.sort(key=os.path.getmtime, reverse=True)
V12_PATH = V12_CANDIDATOS[0] if V12_CANDIDATOS else os.path.join(BASE_DIR, "..", "estoque_consolidada V11.xlsx")

# ws_estoque (ESTOQUE SENIOR) ainda é usado pra montar o catálogo geral (produtos fora do PSI),
# independente do consolidado. Por isso continua referenciando o Sênior direto.
ws_estoque = wb["ESTOQUE SENIOR"]
estoque_por_codigo = {}
estoque_por_nome   = {}
estoque_por_cod_nome = {}  # (cod, nome_normalizado) → qtd. Usado quando o mesmo
                           # cod tem produtos diferentes por empresa (regra 02/07).

def _norm_nome_estoque(s):
    n = ''.join(c if c.isalnum() else ' ' for c in str(s or '').upper())
    return ' '.join(n.split())

# ── FONTE PRIORITÁRIA: posicao_estoque_todas_YYYY-MM-DD.xlsx ───────────────
# Depósitos EXCLUÍDOS (não entram no saldo disponível pro PSI):
#   - RMA (produtos avariados / em análise)
#   - DESCARTES / BAIXA (produtos descartados)
# Todo o resto (CD PORTO BELO, CD ML FULL, CD AMAZON FULL, PRODUTOS/MARKETING, etc)
# entra normalmente.
def _norm_hdr(s):
    return ''.join(c for c in (s or '').upper() if c.isalnum())

if POSICAO_TODAS_PATH and os.path.exists(POSICAO_TODAS_PATH):
    print(f"  Lendo estoque de: {os.path.basename(POSICAO_TODAS_PATH)}")
    wb_pt = openpyxl.load_workbook(POSICAO_TODAS_PATH, data_only=True)
    ws_pt = wb_pt.active
    # Detecta colunas via header (resiliente a reordenação)
    _hdrs_pt = {_norm_hdr(ws_pt.cell(row=1, column=c).value): c
                for c in range(1, ws_pt.max_column + 1)}
    def _col(*names):
        for n in names:
            if _norm_hdr(n) in _hdrs_pt: return _hdrs_pt[_norm_hdr(n)]
        return None
    C_COD      = _col('Produto', 'Código', 'Codigo') or 1
    C_DESC     = _col('Descrição', 'Descricao') or 2
    C_EMP      = _col('Empresa') or 5
    C_DEP      = _col('Depósito', 'Deposito') or 7
    C_FIS      = _col('Físico', 'Fisico') or 9
    C_DISP     = _col('Disponível', 'Disponivel') or 10
    C_RES      = _col('Reservado', 'Reserva')   # opcional
    C_FAM      = _col('Família', 'Familia', 'Grupo')  # opcional
    C_MARCA    = _col('Marca', 'Fabricante')  # opcional

    rma_excl = 0; skus = 0; qtd_total = 0
    por_empresa = {}
    # linhas_estoque: uma linha por (cod, empresa, deposito) — usado pela aba Estoque do frontend.
    # Contém RMA/Descartes/Marketing também (com flag rma=True) pra o usuário filtrar.
    linhas_estoque = []
    for r in range(2, ws_pt.max_row + 1):
        cod  = ws_pt.cell(row=r, column=C_COD).value
        desc = ws_pt.cell(row=r, column=C_DESC).value
        emp  = ws_pt.cell(row=r, column=C_EMP).value
        dep  = ws_pt.cell(row=r, column=C_DEP).value
        fis  = ws_pt.cell(row=r, column=C_FIS).value or 0
        disp = ws_pt.cell(row=r, column=C_DISP).value or 0 if C_DISP else 0
        res_v = ws_pt.cell(row=r, column=C_RES).value or 0 if C_RES else max(0, (fis or 0) - (disp or 0))
        fam  = ws_pt.cell(row=r, column=C_FAM).value if C_FAM else None
        mar  = ws_pt.cell(row=r, column=C_MARCA).value if C_MARCA else None
        if cod is None or not isinstance(fis, (int, float)) or fis <= 0:
            continue
        # Exclui RMA, DESCARTES/BAIXA e MARKETING (não é estoque de venda)
        dep_u = str(dep or '').upper()
        is_rma_ou_excluido = 'RMA' in dep_u or 'DESCARTE' in dep_u or 'BAIXA' in dep_u or 'MARKETING' in dep_u
        try: codigo_original = int(cod)
        except (ValueError, TypeError): codigo_original = cod
        # `chave` = codigo pos-merge (usado na agregacao do PSI — RALWC IMP soma
        # com RALWC principal). `codigo_original` = codigo do cadastro no Senior
        # (usado na aba Estoque — mantem cadastros separados visualmente).
        chave = CODIGOS_MESCLADOS.get(codigo_original, codigo_original)

        # Adiciona à lista de linhas do estoque (visão da aba Estoque)
        emp_lbl = str(emp or '').split(' - ')[-1].strip()
        linhas_estoque.append({
            'codigo':     codigo_original,
            'nome':       str(desc or '').strip(),
            'familia':    str(fam or '').strip() if fam else '',
            'marca':      str(mar or '').strip() if mar else '',
            'empresa':    emp_lbl,
            'deposito':   str(dep or '').strip(),
            'fisico':     int(fis),
            'reservado':  int(res_v) if isinstance(res_v, (int, float)) else 0,
            'disponivel': int(disp) if isinstance(disp, (int, float)) else int(fis),
            'rma':        is_rma_ou_excluido,
        })

        # Agrega no saldo total (só se não é RMA/excluído)
        if is_rma_ou_excluido:
            rma_excl += int(fis); continue
        estoque_por_codigo[chave] = estoque_por_codigo.get(chave, 0) + int(fis)
        if desc:
            nome_upper = str(desc).strip().upper()
            # Normaliza nome do IMP pra bater com o cadastro principal
            nome_upper_merge = nome_upper.replace(' - IMP', '').strip()
            estoque_por_nome[nome_upper_merge] = estoque_por_nome.get(nome_upper_merge, 0) + int(fis)
            # Índice (cod, nome_norm) — quando o mesmo cod tem 2 produtos por empresa,
            # cada linha do estoque vai pro cadastro certo pelo nome normalizado.
            chave_cn = (chave, _norm_nome_estoque(nome_upper_merge))
            estoque_por_cod_nome[chave_cn] = estoque_por_cod_nome.get(chave_cn, 0) + int(fis)
        skus += 1
        qtd_total += int(fis)
        por_empresa[emp_lbl] = por_empresa.get(emp_lbl, 0) + int(fis)

    for emp_lbl, q in sorted(por_empresa.items()):
        print(f"    {emp_lbl}: {q:,} un".replace(',', '.'))
    print(f"  TOTAL: {len(estoque_por_codigo)} SKUs, {qtd_total:,} un ({rma_excl} un de RMA/Descartes excluídas)".replace(',', '.'))
    _usou_novo = True
else:
    _usou_novo = False

# ── FONTE ANTIGA (fallback quando não achou posicao_estoque_todas) ─────────
if not _usou_novo and os.path.exists(V12_PATH):
    print(f"  Lendo estoque consolidado de: {os.path.basename(V12_PATH)}")
    wb_v = openpyxl.load_workbook(V12_PATH, data_only=True)

    # Aba 1: Velds CONSOLIDADO
    # Colunas V12: A=Desc B=Cod OMIE C=Cod Sênior D=Cod Full E=Qtd Omie F=Qtd VELDS
    #              G=Qtd Full(ML) H=Qtd RMA I=Qtd TOTAL J=CMC K=CMV L=Status
    # ⚠ Pra "saldo disponível pro CloudIA", EXCLUI Qtd RMA do total.
    ws_v = wb_v["Velds CONSOLIDADO"]
    headers = [(ws_v.cell(row=1, column=c).value or '').strip() if isinstance(ws_v.cell(row=1, column=c).value, str) else None for c in range(1, ws_v.max_column+1)]
    # Detecta a coluna "Qtd RMA" dinamicamente (caso futuras versões mudem a ordem)
    col_rma = next((i+1 for i, h in enumerate(headers) if h and h.upper() == 'QTD RMA'), 8)
    col_total = next((i+1 for i, h in enumerate(headers) if h and h.upper() == 'QTD TOTAL'), 9)
    col_senior = next((i+1 for i, h in enumerate(headers) if h and 'SÊNIOR' in h.upper()), 3)
    col_desc = 1
    v_skus = 0; v_qtd = 0; v_rma_excluido = 0
    for r in range(2, ws_v.max_row + 1):
        desc = ws_v.cell(row=r, column=col_desc).value
        # ⚠ STOP: separador "EM TRANSITO" (ou variações) marca início do bloco de pedidos
        # em trânsito que NÃO devem entrar no estoque atual.
        if isinstance(desc, str) and 'TRANSITO' in desc.upper().replace('Â', 'A').replace('Ã', 'A'):
            break
        cod_senior_raw = str(ws_v.cell(row=r, column=col_senior).value or '').strip()
        qtd_total = ws_v.cell(row=r, column=col_total).value or 0
        qtd_rma   = ws_v.cell(row=r, column=col_rma).value or 0
        if not isinstance(qtd_total, (int, float)):
            continue
        if not isinstance(qtd_rma, (int, float)):
            qtd_rma = 0
        # Saldo disponível = total - RMA (exclui RMA conforme regra do usuário)
        qtd_disponivel = int(qtd_total) - int(qtd_rma)
        if qtd_disponivel <= 0:
            continue
        v_rma_excluido += int(qtd_rma)
        # Cod Sênior pode ser "400500003 / 400500002" — atribui ao primeiro (canônico)
        cod_principal = cod_senior_raw.split(' / ')[0].strip()
        if cod_principal in ('', '—'):
            continue
        try:
            chave = int(cod_principal)
        except ValueError:
            chave = cod_principal
        estoque_por_codigo[chave] = estoque_por_codigo.get(chave, 0) + qtd_disponivel
        if desc:
            estoque_por_nome[str(desc).strip().upper()] = qtd_disponivel
        v_skus += 1
        v_qtd  += qtd_disponivel

    # Aba 2: Estoque FX - TELECOM
    # Colunas V12: A=Produto B=Desc C=Marca D=Empresa E=Filial F=Depósito G=Qtd H=Custo I=Valor
    # ⚠ Pula linhas onde Depósito contém "RMA"
    ws_fx = wb_v["Estoque FX - TELECOM"]
    fx_skus_set = set(); fx_qtd = 0; fx_rma_excluido = 0
    for row in ws_fx.iter_rows(min_row=2, values_only=True):
        if row[0] is None: continue
        cod  = row[0]
        desc = row[1]
        dep  = row[5]
        qtd  = row[6] or 0
        # ⚠ STOP: separador "EM TRANSITO" marca itens em trânsito (não somar no estoque atual).
        # O texto pode estar em qualquer coluna do primeiro bloco; verifica cod (col 1) e desc (col 2).
        for v in (cod, desc):
            if isinstance(v, str) and 'TRANSITO' in v.upper().replace('Â', 'A').replace('Ã', 'A'):
                qtd = None  # força pular (e qualquer linha depois também via flag)
                break
        if qtd is None:
            break
        if not isinstance(qtd, (int, float)) or qtd == 0:
            continue
        # Exclui depósitos RMA
        if dep and 'RMA' in str(dep).upper():
            fx_rma_excluido += int(qtd)
            continue
        try:
            chave = int(cod) if not isinstance(cod, int) else cod
        except (ValueError, TypeError):
            chave = cod
        estoque_por_codigo[chave] = estoque_por_codigo.get(chave, 0) + int(qtd)
        if desc:
            estoque_por_nome.setdefault(str(desc).strip().upper(), 0)
            estoque_por_nome[str(desc).strip().upper()] += int(qtd)
        fx_skus_set.add(chave)
        fx_qtd += int(qtd)

    print(f"  Velds: {v_skus} SKUs · {v_qtd:,} un disponíveis ({v_rma_excluido} un de RMA excluídas)".replace(',', '.'))
    print(f"  FX-Telecom: {len(fx_skus_set)} SKUs · {fx_qtd:,} un disponíveis ({fx_rma_excluido} un de RMA excluídas)".replace(',', '.'))
    print(f"  TOTAL: {len(estoque_por_codigo)} SKUs no estoque_por_codigo")
elif not _usou_novo:
    # Fallback FINAL: lê direto do PSI_integrado (modo antigo) — só quando NENHUM
    # dos consolidados (posicao_estoque_todas OU Estoque FIBERX) foi encontrado
    print(f"  [AVISO] Consolidado não encontrado — usando ESTOQUE SENIOR direto")
    rma_qtd = 0
    for row in ws_estoque.iter_rows(min_row=2, values_only=True):
        codigo = row[4]
        qtd = row[9]
        deposito = row[8]
        if codigo is None:
            continue
        qtd = int(qtd) if isinstance(qtd, (int, float)) else 0
        if deposito and "RMA" in str(deposito).upper():
            rma_qtd += qtd
            continue
        estoque_por_codigo[codigo] = estoque_por_codigo.get(codigo, 0) + qtd
    for row in ws_estoque.iter_rows(min_row=2, values_only=True):
        nome = row[5]
        qtd  = row[9]
        deposito = row[8]
        if deposito and "RMA" in str(deposito).upper():
            continue
        if nome and qtd:
            estoque_por_nome[nome.strip().upper()] = estoque_por_nome.get(nome.strip().upper(), 0) + int(qtd)
    print(f"  Fallback ESTOQUE SENIOR: {len(estoque_por_codigo)} SKUs (RMA excluído: {rma_qtd} un)")

# ---------------------------------------------------------------------------
# 3.5 ESTOQUE OMIE — só roda no modo fallback (V11 já agrega OMIE+VELDS+Full).
# ---------------------------------------------------------------------------
def omie_para_senior(c):
    """Converte código OMIE pro padrão Sênior inserindo '00' depois dos 4 primeiros chars."""
    s = str(c)
    return (s[:4] + '00' + s[4:]) if len(s) >= 4 else s

import re

def extrair_identificador_vds(desc):
    """Extrai o código VDS da descrição (ex: 'VDS APBW', 'VDS RALW-C', 'VDS CEW-3M').
    Normaliza removendo hifens pra facilitar match."""
    if not desc: return None
    m = re.search(r'VDS\s+([A-Z0-9\-]+)', str(desc).upper())
    if not m: return None
    return m.group(1).replace('-', '')  # normaliza: RALW-C → RALWC

if os.path.exists(V12_PATH):
    pass  # Consolidado ativo: OMIE já está dentro de Qtd TOTAL da aba Velds CONSOLIDADO
else:
    try:
        ws_omie = wb["ESTOQUE OMIE"]
        # Cria índice de identificadores VDS → código Sênior (a partir dos produtos no Sênior).
        # IGNORA linhas cujo nome começa com "PECA " — várias peças compartilham o mesmo
        # VDS-id do produto-pai (ex: "PECA FILTRO ... VDS RALWC" vs "ROBO ASPIRADOR ... VDS RALWC").
        senior_por_vds = {}
        for row in ws_estoque.iter_rows(min_row=2, values_only=True):
            cod_sr = row[4]
            nome_sr = row[5]
            if not nome_sr or str(nome_sr).strip().upper().startswith("PECA "):
                continue
            ident = extrair_identificador_vds(nome_sr)
            if ident and cod_sr is not None:
                senior_por_vds.setdefault(ident, cod_sr)

        omie_qtd_total = 0
        omie_match_count = 0
        omie_match_por_vds = 0
        omie_sem_match = []
        for row in ws_omie.iter_rows(min_row=2, values_only=True):
            cod_omie = row[1]
            desc     = row[0]
            qtd      = row[3]
            if cod_omie is None or not isinstance(qtd, (int, float)) or qtd == 0:
                continue
            qtd_i = int(qtd)
            chave_match = None

            cod_senior_padded = omie_para_senior(cod_omie)
            try:
                cod_senior_int = int(cod_senior_padded)
            except (ValueError, TypeError):
                cod_senior_int = None
            if cod_senior_int is not None and cod_senior_int in estoque_por_codigo:
                chave_match = cod_senior_int
            elif cod_senior_padded in estoque_por_codigo:
                chave_match = cod_senior_padded

            if chave_match is None:
                ident = extrair_identificador_vds(desc)
                if ident and ident in senior_por_vds:
                    chave_match = senior_por_vds[ident]
                    omie_match_por_vds += 1

            if chave_match is not None:
                estoque_por_codigo[chave_match] += qtd_i
                omie_qtd_total += qtd_i
                omie_match_count += 1
            else:
                omie_sem_match.append((str(desc)[:50], cod_omie, qtd_i))

        print(f"  {omie_qtd_total} unidades OMIE somadas ao Sênior ({omie_match_count} entradas · {omie_match_por_vds} via VDS-id)")
        if omie_sem_match:
            print(f"  ({len(omie_sem_match)} linhas OMIE sem match — ignorados)")
    except KeyError:
        print("  (aba ESTOQUE OMIE não encontrada — pulando)")

# ---------------------------------------------------------------------------
# 4. Kanban (pedidos_kanban.json) — fonte única de chegadas desde 26/06/2026
# Religado em 26/06/2026: passa a popular sell_in semanal + entradas (prox_chegada)
# do PSI a partir do pedidos_kanban.json. Antes disso (02/06–26/06), o sell_in
# vinha SÓ da planilha PSI Excel — agora é "kanban com fallback pra Excel".
#
# Regra de combinação por (produto, semana):
#   - Se kanban_sin > 0  → usa kanban (Excel é assumido como obsoleto)
#   - Se kanban_sin == 0 e Excel > 0 → mantém Excel (entrada antiga preservada)
# Discrepâncias (kanban>0 e Excel>0 com valor diferente) são logadas pra auditoria.
#
# Pedidos cancelados (`cancelado: true`) são SEMPRE ignorados.
# Todas as fases do kanban contam (pendente até estoque) — porque cada pedido tem
# uma `semana` de chegada planejada que precisa entrar como sell_in dessa semana.
# ---------------------------------------------------------------------------
PEDIDOS_KANBAN_PATH = os.path.join(os.path.dirname(__file__), "pedidos_kanban.json")

kanban_sin_por_nome = {}        # nome_upper → {semana: qtd_total}
kanban_sin_por_codigo = {}      # codigo (int) → {semana: qtd_total}
kanban_entradas_por_nome = {}   # nome_upper → [{semana, quantidade, status, mes, codigo_pedido, fase}]
kanban_entradas_por_codigo = {} # codigo (int) → [...]
# Sell_in EFETIVO: só pedidos que fisicamente chegaram (fase estoque/recebido).
# Usado no cascade backward de W-1 pra distinguir "chegou de verdade" (foto já reflete)
# de "planejado, mas ainda em produção/trânsito/numerário" (foto NÃO reflete).
kanban_sin_efetivo_por_nome = {}
kanban_sin_efetivo_por_codigo = {}

# Retorna a fase real do Kanban em uppercase (ex: PENDENTE, PAGO1, PRODUCAO,
# INSPECAO, PAGO2, EMBARQUE, PORTO, NUMERARIO, TRANSITO, RECEBIDO, ESTOQUE).
# Antes agrupava tudo em 3 buckets (PEDIDO / EM_TRANSITO / RECEBIDO), o que
# ficava enganoso: um pedido em EMBARQUE virava "PEDIDO" no dados.json, e
# quem lia depois achava que ainda estava em fase 'pendente'. Douglas
# apontou 31/08/2026. O calculo de `in_transit` foi migrado pra usar
# entrada['fase'] direto (embarque/porto/numerario/transito).
def _fase_para_status(fase: str) -> str:
    return (fase or "pendente").upper()

try:
    with open(PEDIDOS_KANBAN_PATH, encoding="utf-8") as _f:
        _pedidos_kanban_raw = json.load(_f)
    _total_kanban = 0
    for _ped in _pedidos_kanban_raw:
        if not isinstance(_ped, dict):
            continue
        if _ped.get("cancelado"):
            continue
        _nome = (_ped.get("nome") or "").strip().upper()
        _semana = _ped.get("semana")
        _qtd = _ped.get("qtd") or 0
        if not _nome or not isinstance(_semana, (int, float)) or not _qtd:
            continue
        _semana = int(_semana)
        _qtd = int(_qtd)
        # Código do produto (Sênior) — opcional. Quando presente, vira o canal
        # primário de match. Permite nomes ligeiramente diferentes entre kanban
        # e PSI sem perder o sell_in.
        _codigo_prod = _ped.get("codigo")
        try:
            _codigo_prod = int(_codigo_prod) if _codigo_prod else None
        except (ValueError, TypeError):
            _codigo_prod = None
        _entrada = {
            "semana":     _semana,
            "quantidade": _qtd,
            "status":     _fase_para_status(_ped.get("fase")),
            "mes":        _ped.get("mes", ""),
            "codigo_pedido": _ped.get("codigo_pedido", ""),
            "fase":       _ped.get("fase", ""),
        }
        # Indexa por NOME (sempre)
        kanban_sin_por_nome.setdefault(_nome, {})
        kanban_sin_por_nome[_nome][_semana] = kanban_sin_por_nome[_nome].get(_semana, 0) + _qtd
        kanban_entradas_por_nome.setdefault(_nome, []).append(_entrada)
        # Indexa também por CÓDIGO quando presente
        if _codigo_prod is not None:
            kanban_sin_por_codigo.setdefault(_codigo_prod, {})
            kanban_sin_por_codigo[_codigo_prod][_semana] = kanban_sin_por_codigo[_codigo_prod].get(_semana, 0) + _qtd
            kanban_entradas_por_codigo.setdefault(_codigo_prod, []).append(_entrada)
        # EFETIVO: só fase 'estoque' — pedido teve NF lançada e o saldo do Sênior
        # ja reflete. 'Recebido' significa chegou no galpao mas AINDA nao subiu
        # no Sênior, entao pra fins de cascade continua contando como planejado.
        _fase_ped = (_ped.get("fase") or "").lower()
        if _fase_ped == "estoque":
            kanban_sin_efetivo_por_nome.setdefault(_nome, {})
            kanban_sin_efetivo_por_nome[_nome][_semana] = kanban_sin_efetivo_por_nome[_nome].get(_semana, 0) + _qtd
            if _codigo_prod is not None:
                kanban_sin_efetivo_por_codigo.setdefault(_codigo_prod, {})
                kanban_sin_efetivo_por_codigo[_codigo_prod][_semana] = kanban_sin_efetivo_por_codigo[_codigo_prod].get(_semana, 0) + _qtd
        _total_kanban += 1
    print(f"  Kanban: {_total_kanban} pedidos ativos lidos de pedidos_kanban.json ({len(kanban_sin_por_nome)} produtos)")
except FileNotFoundError:
    print(f"  AVISO: pedidos_kanban.json não encontrado em {PEDIDOS_KANBAN_PATH} — sell_in virá só do Excel")

# entradas_por_codigo é o índice usado pelo loop de produtos (linha ~582 abaixo).
# Indexa simultaneamente por código E por nome — a busca lá embaixo tenta código
# primeiro, depois nome (fallback). Pedido com `codigo` no kanban encontra match
# mesmo se o nome divergir do PSI.
entradas_por_codigo = {}
for _cod, _lista in kanban_entradas_por_codigo.items():
    entradas_por_codigo[_cod] = sorted(_lista, key=lambda e: e["semana"])
for _nome_upper, _lista in kanban_entradas_por_nome.items():
    # Não sobrescreve se já indexado por código
    if _nome_upper not in entradas_por_codigo:
        entradas_por_codigo[_nome_upper] = sorted(_lista, key=lambda e: e["semana"])

# ---------------------------------------------------------------------------
# 5. cenario 1 — séries semanais por produto
# ---------------------------------------------------------------------------
ws_c1 = wb["cenario 1"]

def ler_psi_produto(linha_psi: int, off_fcst: int, off_sin: int, off_real: int, off_sld: int):
    """
    Lê as séries semanais de um bloco PSI.
    Os offsets da aba Lista são relativos ao cabeçalho do produto (linha_psi - 1).
    A linha de semanas fica em linha_psi, e os dados em (linha_psi - 1) + offset.
    """
    base = linha_psi - 1  # linha do cabeçalho do produto (nome, marca, meses)
    row_semanas = [ws_c1.cell(row=linha_psi,        column=c).value for c in range(3, 200)]
    row_fcst    = [ws_c1.cell(row=base + off_fcst,  column=c).value for c in range(3, 200)]
    row_sin     = [ws_c1.cell(row=base + off_sin,   column=c).value for c in range(3, 200)]
    row_real    = [ws_c1.cell(row=base + off_real,  column=c).value for c in range(3, 200)]
    row_saldo   = [ws_c1.cell(row=base + off_sld,   column=c).value for c in range(3, 200)]

    # Corta após colunas vazias (sem rótulo de semana)
    semanas_dict = {}  # semana_num → {fcst, sin, real, saldo}
    for i, label in enumerate(row_semanas):
        if label is None:
            break
        sem = parse_week_label(label)
        if sem is None:
            continue
        def n(v):
            return float(v) if isinstance(v, (int, float)) else 0.0
        if sem not in semanas_dict:
            semanas_dict[sem] = {"fcst": 0.0, "sin": 0.0, "real": 0.0, "saldo": 0.0}
        # Semanas-fronteira aparecem em dois meses: usa a última ocorrência (mês seguinte vence)
        if n(row_fcst[i]) != 0:
            semanas_dict[sem]["fcst"] = n(row_fcst[i])
        if n(row_sin[i]) != 0:
            semanas_dict[sem]["sin"]  = n(row_sin[i])
        if n(row_real[i]) != 0:
            semanas_dict[sem]["real"] = n(row_real[i])
        if n(row_saldo[i]) != 0:
            semanas_dict[sem]["saldo"] = n(row_saldo[i])

    semanas_ord = sorted(semanas_dict.keys())
    return {
        "semanas":   semanas_ord,
        "forecast":  [round(semanas_dict[s]["fcst"])  for s in semanas_ord],
        "sell_in":   [round(semanas_dict[s]["sin"])   for s in semanas_ord],
        "realizado": [round(semanas_dict[s]["real"])  for s in semanas_ord],
        "saldo":     [round(semanas_dict[s]["saldo"]) for s in semanas_ord],
    }

# ---------------------------------------------------------------------------
# 5.5 VENDAS B2B — agregadas por (produto, semana, mes calendário)
# Cada venda traz uma data; semana é derivada via semana_de_data e o "mês PSI" é
# o mês CALENDÁRIO da data (não o mês-dono da semana). Isso permite, em semanas-
# fronteira (W23, W27, W31...), distinguir vendas feitas em JUN vs JUL no mesmo W27.
# ---------------------------------------------------------------------------
MESES_ABR = ['JAN','FEV','MAR','ABR','MAI','JUN','JUL','AGO','SET','OUT','NOV','DEZ']

def mes_de_data(data):
    """Converte data → 'MAI/26', baseado no mês calendário (não no mês-dono PSI)."""
    if isinstance(data, datetime.datetime):
        data = data.date()
    if not isinstance(data, datetime.date):
        return None
    return f"{MESES_ABR[data.month - 1]}/{(data.year - 2000):02d}"

# Fonte preferida: arquivo "vendas_b2b_<data>.xlsx" na pasta Posição estoques
# (relatório exportado direto do sistema). Se não existir, cai pra aba VENDAS B2B
# do PSI_integrado (modo antigo). Detecção de colunas é feita por header pra
# tolerar pequenas variações de layout entre os dois.
def _norm_header(h):
    if h is None:
        return ""
    s = str(h).strip().lower()
    # remove acentos básicos pra casar "qtde" / "qtd"
    return (s.replace("ã", "a").replace("á", "a").replace("â", "a")
              .replace("é", "e").replace("ê", "e").replace("í", "i")
              .replace("ó", "o").replace("ô", "o").replace("õ", "o")
              .replace("ú", "u").replace("ç", "c"))

def _achar_idx(headers_norm, *candidatos):
    for cand in candidatos:
        cand_n = _norm_header(cand)
        for i, h in enumerate(headers_norm):
            if h == cand_n:
                return i
        # match parcial (header começa com candidato)
        for i, h in enumerate(headers_norm):
            if h.startswith(cand_n):
                return i
    return None

# Procura em "Cloud PCM/VENDAS/" (preferida — relatório direto do sistema) e em
# "Cloud PCM/Posição estoques/" (alternativa). Aceita "VENDAS B2B 02.06.xlsx",
# "vendas_b2b_<data>.xlsx" etc. — qualquer .xlsx com "vendas b2b" / "vendas_b2b" no nome.
_PASTA_VENDAS = os.path.join(BASE_DIR, "..", "VENDAS")
_B2B_CANDIDATOS = []
for _pasta in (_PASTA_VENDAS, _PASTA_POSICAO):
    if os.path.isdir(_pasta):
        for f in os.listdir(_pasta):
            fl = f.lower()
            if fl.endswith(".xlsx") and ("vendas_b2b" in fl or "vendas b2b" in fl):
                _B2B_CANDIDATOS.append(os.path.join(_pasta, f))
_B2B_CANDIDATOS.sort(key=os.path.getmtime, reverse=True)

if _B2B_CANDIDATOS:
    _B2B_PATH = _B2B_CANDIDATOS[0]
    print(f"  Lendo vendas B2B de: {os.path.basename(_B2B_PATH)}")
    _wb_b2b = openpyxl.load_workbook(_B2B_PATH, data_only=True, read_only=True)
    ws_b2b = _wb_b2b.active
else:
    print(f"  [AVISO] VENDAS B2B *.xlsx não encontrado em VENDAS/ ou Posição estoques/ — usando aba VENDAS B2B do PSI")
    ws_b2b = wb["VENDAS B2B"]

# Mapeia colunas via header
_headers_raw = next(ws_b2b.iter_rows(min_row=1, max_row=1, values_only=True))
_headers = [_norm_header(h) for h in _headers_raw]

idx_data    = _achar_idx(_headers, "data")
idx_nf      = _achar_idx(_headers, "nf", "numero nf", "n nf")
idx_cliente = _achar_idx(_headers, "cliente", "razao social", "razao")
idx_uf      = _achar_idx(_headers, "uf", "estado")
idx_cod     = _achar_idx(_headers, "cod", "codigo", "sku")
idx_nome    = _achar_idx(_headers, "material", "produto", "descricao")
idx_marca   = _achar_idx(_headers, "marca")
idx_qtd     = _achar_idx(_headers, "qtde", "qtd", "quantidade")
idx_valor   = _achar_idx(_headers, "faturamento", "valor", "valor total", "total")

if idx_data is None or idx_qtd is None:
    raise RuntimeError(f"Cabeçalho de vendas B2B inválido — não achei Data/Qtd. Headers: {_headers_raw}")

vendas_b2b_codigo = {}   # codigo → {"semana_mes": qtd_acumulada}
vendas_b2b_nome   = {}   # nome upper → {"semana_mes": qtd_acumulada}
vendas_b2b_linhas = []   # linhas brutas pra alimentar a aba Vendas no front

for row in ws_b2b.iter_rows(min_row=2, values_only=True):
    data_v = row[idx_data] if idx_data is not None else None
    codigo = row[idx_cod]  if idx_cod  is not None else None
    nome_v = row[idx_nome] if idx_nome is not None else None
    qtd    = row[idx_qtd]  if idx_qtd  is not None else None
    if not data_v or not qtd:
        continue
    try:
        qtd_i = int(qtd)
    except (ValueError, TypeError):
        continue
    sem = semana_de_data(data_v)
    mes_cal = mes_de_data(data_v)
    if not mes_cal or sem < 1:
        continue
    key = f"{sem}_{mes_cal}"
    if codigo:
        # Normaliza: VENDAS vem com cod string, Lista PSI com int — converte tudo pra str.
        cod_norm = str(codigo).strip()
        d_ = vendas_b2b_codigo.setdefault(cod_norm, {})
        d_[key] = d_.get(key, 0) + qtd_i
    if nome_v:
        nm = str(nome_v).strip().upper()
        d_ = vendas_b2b_nome.setdefault(nm, {})
        d_[key] = d_.get(key, 0) + qtd_i

    # Linha bruta pro front (Vendas tab)
    nf_v     = row[idx_nf]     if idx_nf     is not None else None
    cli_v    = row[idx_cliente] if idx_cliente is not None else None
    uf_v     = row[idx_uf]     if idx_uf     is not None else None
    marca_v  = row[idx_marca]  if idx_marca  is not None else None
    valor_v  = row[idx_valor]  if idx_valor  is not None else None
    try:
        valor_f = float(valor_v) if valor_v is not None else 0.0
    except (ValueError, TypeError):
        valor_f = 0.0
    vendas_b2b_linhas.append({
        "data":    data_v.isoformat() if isinstance(data_v, (datetime.date, datetime.datetime)) else str(data_v),
        "nf":      str(nf_v) if nf_v is not None else "",
        "cliente": str(cli_v).strip() if cli_v else "",
        "uf":      str(uf_v).strip() if uf_v else "",
        "codigo":  codigo,
        "produto": str(nome_v).strip() if nome_v else "",
        "marca":   str(marca_v).strip() if marca_v else "",
        "qtd":     qtd_i,
        "valor":   round(valor_f, 2),
        "semana":  sem,
        "mes":     mes_cal,
    })

# Ordena por data desc pra UI mostrar lançamentos mais recentes primeiro
vendas_b2b_linhas.sort(key=lambda x: x["data"], reverse=True)

total_b2b = sum(sum(v.values()) for v in vendas_b2b_codigo.values())
print(f"  {total_b2b} unidades em VENDAS B2B agregadas por (semana, mês) — {len(vendas_b2b_linhas)} linhas")

# ---------------------------------------------------------------------------
# 6. Montar lista final de produtos
# ---------------------------------------------------------------------------
MESES_FCST = list(WEEK_TO_MONTH.values())  # lista de meses na ordem

# Override de código pra produtos PSI cujo Excel ainda não tem coluna de código
# preenchida (ex: V30 só foi cadastrado no Sênior depois do PSI). Match por
# nome upper-strip. Adicionar entradas conforme novos produtos forem cadastrados
# no Sênior. Sem isso, o autosync kanban→PSI desses produtos só funciona por
# nome exato (frágil).
PSI_CODIGO_OVERRIDE = {
    "ROBO ASPIRADOR V30": 400700025,  # cadastrado em jun/2026
}

# Histórico: V50/V50PRO foram recadastrados no Sênior (Omie→Sênior) com códigos
# novos longos (400700012-014). Chegadas novas já entram sob código novo, então
# não precisamos mais do fallback de saldo. Mantido zerado por compatibilidade.
_MIGRACAO_CODIGO_REV = {}

produtos_final = []
for p in produtos_raw:
    codigo = p["codigo"]
    nome   = p["nome"]
    if codigo is None:
        codigo = PSI_CODIGO_OVERRIDE.get((nome or "").strip().upper())

    # Estoque atual
    # Exceções (Alimentador/Robô): foto do Sênior não é fonte de verdade — usa o saldo
    # congelado da semana anterior à SEMANA_ESTOQUE (vem do histórico) e cascade segue dali.
    if codigo in EXCECOES_FOTO_SENIOR:
        hist_prod = HISTORICO_SALDO.get(str(codigo), {})
        # saldo final da semana imediatamente anterior à SEMANA_ESTOQUE = inicial da SEMANA_ESTOQUE
        saldo = int(hist_prod.get(str(SEMANA_ESTOQUE - 1), 0))
    else:
        saldo = estoque_por_codigo.get(codigo, 0)
        if saldo == 0:
            saldo = estoque_por_nome.get(nome.strip().upper(), 0)

    # Chegadas
    entradas = entradas_por_codigo.get(codigo, [])
    if not entradas:
        entradas = entradas_por_codigo.get(nome.strip().upper(), [])
    entradas_futuras = sorted(
        [e for e in entradas if e["semana"] >= SEMANA_ATUAL],
        key=lambda e: e["semana"]
    )
    # in_transit = pedidos que ja sairam do fornecedor e estao chegando (embarque
    # em diante). Antes usava status agrupado 'EM_TRANSITO'; agora le fase direto.
    _FASES_EM_TRANSITO = {"embarque", "porto", "numerario", "transito"}
    in_transit = sum(e["quantidade"] for e in entradas_futuras
                     if (e.get("fase") or "").lower() in _FASES_EM_TRANSITO)
    prox_chegada = entradas_futuras[0] if entradas_futuras else None

    # Forecast semanal médio: total dos 3 primeiros meses ÷ número real de semanas
    MESES_ORDEM = [
        'MAI/26','JUN/26','JUL/26','AGO/26','SET/26','OUT/26',
        'NOV/26','DEZ/26','JAN/27','FEV/27','MAR/27','ABR/27','MAI/27',
    ]
    meses_com_fc = [m for m in MESES_ORDEM if p["fcst_mensal"].get(m, 0) > 0][:3]
    if meses_com_fc:
        total_fc  = sum(p["fcst_mensal"].get(m, 0) for m in meses_com_fc)
        total_sem = sum(semanas_do_mes(m) for m in meses_com_fc)
        fc_sem = round(total_fc / total_sem) if total_sem > 0 else 0
    else:
        fc_sem = 0

    # Tipo
    tipo = tipo_por_produto.get(nome, "Nacional")

    # PSI semanal
    psi = None
    if p["_linha_psi"]:
        try:
            psi = ler_psi_produto(
                int(p["_linha_psi"]),
                int(p["_off_fcst"]),
                int(p["_off_sin"]),
                int(p["_off_real"]),
                int(p["_off_sld"]),
            )
        except Exception as e:
            print(f"  AVISO: erro ao ler PSI de '{nome}': {e}")

    # Aplica realizados AUTOMÁTICOS via VENDAS B2B (agrega por semana) no psi.realizado.
    # Antes: psi.realizado vinha só do PSI Excel (linha _real), que ficava desatualizado
    # em relação ao VENDAS B2B que o V3 exibe no display. Isso causava cascade errado.
    # Agora: soma VENDAS B2B por semana e sobrescreve. REALIZADO_MANUAL sobrescreve depois
    # (pra casos onde B2B não pega, tipo Alimentador/RALWC que vendem via ML).
    if psi:
        vendas_split = vendas_b2b_codigo.get(str(codigo).strip(), {}) or \
                       vendas_b2b_nome.get(nome.strip().upper(), {})
        if vendas_split:
            # Agrupa split "sem_mes" por semana (soma todas as instâncias da semana)
            por_semana_b2b = {}
            for key, qtd in vendas_split.items():
                try:
                    sem_int = int(key.split("_")[0])
                except (ValueError, IndexError):
                    continue
                por_semana_b2b[sem_int] = por_semana_b2b.get(sem_int, 0) + int(qtd)
            for sem_alvo, qtd in por_semana_b2b.items():
                if sem_alvo in psi["semanas"]:
                    psi["realizado"][psi["semanas"].index(sem_alvo)] = int(qtd)

    # Aplica realizados manuais no psi.realizado (sobrescreve B2B — casos ML/etc)
    if psi and codigo in REALIZADO_MANUAL:
        for sem_alvo, qtd in REALIZADO_MANUAL[codigo].items():
            if sem_alvo in psi["semanas"]:
                idx_real = psi["semanas"].index(sem_alvo)
                psi["realizado"][idx_real] = int(qtd)

    # Sobrescreve sell_in semanal com kanban. Kanban é a fonte ÚNICA de verdade
    # (decisão Douglas, 26/06/2026): semana sem pedido no kanban → sell_in=0,
    # independente do que o Excel diga. Pedidos antigos no Excel sem
    # correspondente no kanban são considerados obsoletos.
    # Match prioriza CÓDIGO (estável) e cai pra NOME (frágil) se kanban não
    # tiver código preenchido pra esse produto.
    if psi:
        sin_kanban_prod = kanban_sin_por_codigo.get(codigo, {})
        if not sin_kanban_prod:
            sin_kanban_prod = kanban_sin_por_nome.get(nome.strip().upper(), {})
        for i, sem in enumerate(psi["semanas"]):
            psi["sell_in"][i] = sin_kanban_prod.get(sem, 0)

    # Aplica saldo histórico (semanas fechadas que já foram congeladas em runs anteriores)
    # floor=0: histórico de runs anteriores ao max(0,...) pode ter negativos — sanitiza aqui
    hist_prod = HISTORICO_SALDO.get(str(codigo), {})
    if psi and hist_prod:
        for sem_str, saldo_hist in hist_prod.items():
            sem_int = int(sem_str)
            if sem_int in psi["semanas"]:
                psi["saldo"][psi["semanas"].index(sem_int)] = max(0, int(saldo_hist))

    # Recalcula saldo PSI a partir de SEMANA_ESTOQUE (foto do Sênior) e cascateia até o fim.
    # Estoque inicial da SEMANA_ESTOQUE = saldo do Sênior (foto). Cascade segue daí.
    # saldo_atual exibido no app = saldo projetado em SEMANA_ATUAL.
    # Regra de consumo:
    #   - Semana fechada (i <= SEMANA_ATUAL) com realizado > 0: usa realizado
    #   - Semana futura ou sem realizado: usa forecast
    if psi and SEMANA_ESTOQUE in psi["semanas"]:
        idx = psi["semanas"].index(SEMANA_ESTOQUE)
        # Saldo INICIAL da SEMANA_ESTOQUE = saldo (foto Sênior). Frontend lê inicial = saldo[idx-1]
        if idx > 0:
            # BUG FIX (13/07/2026): se a semana anterior a SEMANA_ESTOQUE é uma semana
            # FECHADA (< SEMANA_ATUAL) com realizado > 0, o SALDO FINAL dela deve vir
            # da conta (saldo[idx-2] + sell_in - realizado), NÃO da foto do Sênior.
            # A foto = Estoque Inicial da SEMANA_ATUAL (que pode diferir do Saldo Final
            # da semana anterior quando há devoluções/ajustes fora do canal de vendas).
            # Sem esse fix, W28 (fechada) mostrava 1036 (foto) em vez de 1019 (conta).
            w_ant = psi["semanas"][idx - 1]
            rl_ant = psi.get("realizado", [0]*len(psi["semanas"]))[idx - 1]
            if w_ant < SEMANA_ATUAL and rl_ant > 0 and idx >= 2:
                saldo_ant_ant = psi["saldo"][idx - 2]
                si_ant = psi["sell_in"][idx - 1]
                psi["saldo"][idx - 1] = max(0, round(saldo_ant_ant + si_ant - rl_ant))
            else:
                psi["saldo"][idx - 1] = max(0, round(saldo))
        def _consumo(i):
            w = psi["semanas"][i]
            fc = psi["forecast"][i]
            rl = psi.get("realizado", [0]*len(psi["semanas"]))[i]
            return rl if (w <= SEMANA_ATUAL and rl > 0) else fc
        # Passado (< SEMANA_ESTOQUE-1): saldo vem do histórico congelado (foto real
        # de rodagens anteriores). Se não tiver histórico, fica 0 — mais honesto
        # que inventar com cascade backward usando forecast (gerava dente-de-serra).
        # CASCADE FORWARD a partir de SEMANA_ESTOQUE.
        psi["saldo"][idx] = max(0, round(saldo - _consumo(idx) + psi["sell_in"][idx]))
        for i in range(idx + 1, len(psi["semanas"])):
            psi["saldo"][i] = max(0, round(psi["saldo"][i-1] + psi["sell_in"][i] - _consumo(i)))
        # saldo_atual exibido = projeção da SEMANA_ATUAL (não o saldo bruto do Sênior)
        if SEMANA_ATUAL in psi["semanas"] and SEMANA_ATUAL != SEMANA_ESTOQUE:
            idx_atual = psi["semanas"].index(SEMANA_ATUAL)
            saldo = max(0, psi["saldo"][idx_atual - 1]) if idx_atual > 0 else saldo

    # Congela no histórico o saldo final de todas as semanas <= SEMANA_ESTOQUE (semanas fechadas)
    # Floor=0: nunca grava saldo negativo no histórico — evita bug onde forecast alto
    # + realizado zerado gera cascade negativo que congelaria valor irreal (ex: -12.315
    # em rodagem passada). Se cascade tenta gravar negativo, fica 0.
    if psi:
        novo_hist = dict(hist_prod)  # preserva entradas antigas
        for w in psi["semanas"]:
            if w <= SEMANA_ESTOQUE:
                saldo_w = psi["saldo"][psi["semanas"].index(w)]
                novo_hist[str(w)] = max(0, int(saldo_w))
        HISTORICO_SALDO[str(codigo)] = novo_hist

    # Realizado split por (semana, mes calendário) — vindo de VENDAS B2B
    # Normaliza pra str porque VENDAS vem string e Lista PSI vem int.
    realizado_split = vendas_b2b_codigo.get(str(codigo).strip(), {})
    if not realizado_split:
        realizado_split = vendas_b2b_nome.get(nome.strip().upper(), {})
    # Aplica realizados manuais no split (display da tabela usa essa estrutura)
    if codigo in REALIZADO_MANUAL:
        realizado_split = dict(realizado_split)  # cópia pra não mutar o original
        for sem_alvo, qtd in REALIZADO_MANUAL[codigo].items():
            mes_alvo = WEEK_TO_MONTH.get(sem_alvo)
            if mes_alvo is None:
                for w, m in sorted(WEEK_TO_MONTH.items()):
                    if w <= sem_alvo: mes_alvo = m
            if mes_alvo:
                realizado_split[f"{sem_alvo}_{mes_alvo}"] = int(qtd)

    produtos_final.append({
        "nome":            nome,
        "marca":           p["marca"],
        "codigo":          codigo,
        "familia":         p["familia"],
        "tipo":            tipo,
        "pv":              p["pv"],
        "pc":              p["pc"],
        "saldo_atual":     saldo,
        "fc_sem":          fc_sem,
        "in_transit":      in_transit,
        "prox_chegada":    prox_chegada,
        "entradas":        entradas_futuras,
        "fcst_mensal":     p["fcst_mensal"],
        "venda_mensal":    p["venda_mensal"],
        "realizado_split": realizado_split,
        "psi":             psi,
    })

print(f"  {len(produtos_final)} produtos montados")

# ---------------------------------------------------------------------------
# 7. Catálogo geral — 436 produtos do ESTOQUE SENIOR (pra busca em Novo Pedido)
# Produtos que não estão na Lista PSI ficam aqui pra criar pedidos sob demanda
# sem aparecer no PSI/Cobertura. Inclui OMIE somado (mesmo critério do saldo dos 65).
# ---------------------------------------------------------------------------
# Padrões de marca detectáveis no nome — extrai marca automaticamente do catálogo.
# Ordem importa: marcas mais específicas primeiro (ex: TP-Link antes de uma genérica "TP").
MARCAS_PADRAO = [
    ('TP-Link',   re.compile(r'\bTP[\s\-]*LINK\b', re.I)),
    ('TP-Link',   re.compile(r'\bMERCUSYS\b', re.I)),  # Mercusys é submarca da TP-Link
    ('Zyxel',     re.compile(r'\bZYXEL\b', re.I)),
    ('Huawei',    re.compile(r'\bHUAWEI\b', re.I)),
    ('Sandisk',   re.compile(r'\bSANDISK\b', re.I)),
    ('Fastt10',   re.compile(r'\bFASTT10\b', re.I)),
    ('Fibratech', re.compile(r'\bFIBRATECH\b', re.I)),
    ('Velds',     re.compile(r'\bVELDS\b|\bVDS\b', re.I)),
]
def detectar_marca(nome):
    if not nome: return ''
    s = str(nome)
    for marca, pat in MARCAS_PADRAO:
        if pat.search(s): return marca
    return ''

codigos_psi = {p["codigo"] for p in produtos_final if p.get("codigo") is not None}
catalogo = []

# Fonte principal do catálogo: produtos_seniorx_<data>.xlsx — exportação completa do Sênior
# (todos os SKUs, mesmo sem estoque). Tem duplicação por CNPJ (1=FX, 3=TELECOM, 7=VELDS) —
# deduplicamos por (código, nome_normalizado): quando o mesmo cod tem nomes iguais em
# empresas diferentes, mantém 1; quando os nomes são realmente distintos (produtos
# diferentes reaproveitando o código), mantém os DOIS. Regra do Douglas 02/07/2026.
# Fallback: se o arquivo não existir, monta do ESTOQUE SENIOR (modo antigo).
import glob
# Busca em 2 lugares: 'Posição estoques' (novo padrão) e pasta pai (legado).
# Ordena por data DESC (nome tem YYYY-MM-DD) e pega o mais recente absoluto.
seniorx_candidatos = sorted(
    glob.glob(os.path.join(BASE_DIR, "..", "Posição estoques", "produtos_seniorx*.xlsx")) +
    glob.glob(os.path.join(BASE_DIR, "..", "produtos_seniorx*.xlsx")),
    key=lambda p: os.path.basename(p),  # nome com data no formato produtos_seniorx_YYYY-MM-DD.xlsx
    reverse=True
)
SENIORX_PATH = seniorx_candidatos[0] if seniorx_candidatos else None

EMPRESA_LABEL = {1: 'FX', 3: 'TELECOM', 7: 'VELDS'}

# Marcas a IGNORAR no catálogo (são categorias internas, não produtos comerciais).
MARCAS_EXCLUIR_CATALOGO = {
    'MATERIAL DE USO E CONSUMO',  # Brindes, custo serviço, materiais de uso interno
}

if SENIORX_PATH:
    print(f"  Catálogo: {os.path.basename(SENIORX_PATH)}")
    wb_sx = openpyxl.load_workbook(SENIORX_PATH, data_only=True)
    ws_sx = wb_sx['Produtos']
    # Detecta colunas pelos headers (Excel velho de 02/06 tem A=Código, novo de 25/06
    # tem A=ID/UUID e B=Código — o layout do export do Sênior mudou).
    def _norm_header(s):
        return ''.join(c for c in (s or '').upper() if c.isalnum())
    headers_sx = {_norm_header(ws_sx.cell(row=1, column=c).value): c
                  for c in range(1, ws_sx.max_column + 1)}
    def _col(*names):
        for n in names:
            if _norm_header(n) in headers_sx:
                return headers_sx[_norm_header(n)]
        return None
    COL_COD     = _col('Código', 'Codigo') or 1
    COL_DESC    = _col('Descrição', 'Descricao') or 2
    COL_REF     = _col('Código Referência', 'Codigo Referencia') or 3
    COL_EMPRESA = _col('Código Empresa', 'Codigo Empresa') or 4
    COL_MARCA   = _col('Marca') or 5
    COL_STATUS  = _col('Status')  # opcional — pula deletados/inativos se existir
    cods_ja = set()  # dedup por (cod, nome_normalizado) — vide comentário no topo
    def _norm_nome(s):
        # Ignora caixa, pontuação e espaços duplos pra tratar "S6730/5731" == "S67305731"
        n = ''.join(c if c.isalnum() else ' ' for c in str(s or '').upper())
        return ' '.join(n.split())
    estoque_por_codigo_str = {str(k): v for k, v in estoque_por_codigo.items()}  # busca por string também
    # Códigos que TÊM discriminação por nome no estoque — usados pra decidir se o
    # fallback pelo cod-puro é seguro (evita dar o total pra 2 produtos diferentes).
    cods_com_discriminacao = {k[0] for k in estoque_por_cod_nome.keys()}
    for r in range(2, ws_sx.max_row + 1):
        cod = ws_sx.cell(row=r, column=COL_COD).value
        desc = ws_sx.cell(row=r, column=COL_DESC).value
        ref = ws_sx.cell(row=r, column=COL_REF).value
        empresa = ws_sx.cell(row=r, column=COL_EMPRESA).value
        marca_excel = ws_sx.cell(row=r, column=COL_MARCA).value
        if COL_STATUS:
            status_prod = ws_sx.cell(row=r, column=COL_STATUS).value
            if status_prod and str(status_prod).strip().lower() not in ('ativo', '-'):
                continue
        if cod is None or not desc:
            continue
        marca_norm = str(marca_excel or '').strip().upper()
        # Filtra categorias internas que não são produtos comerciais (brindes, etc.)
        if marca_norm in MARCAS_EXCLUIR_CATALOGO:
            continue
        # Normaliza pra int quando possível
        try: cod_int = int(cod)
        except (ValueError, TypeError): cod_int = cod
        # Pula se já está no PSI (PSI tem prioridade — produto não aparece duas vezes)
        if cod_int in codigos_psi:
            continue
        # Dedup por (cod, nome_normalizado): duplicação pura por CNPJ some,
        # mas dois produtos diferentes reaproveitando o mesmo código sobrevivem.
        chave_dedup = (cod_int, _norm_nome(desc))
        if chave_dedup in cods_ja:
            continue
        cods_ja.add(chave_dedup)
        # Marca: usa do Excel se vier preenchida e for marca conhecida, senão detecta da desc
        marca_map = {
            'TP-LINK': 'TP-Link', 'TPLINK': 'TP-Link', 'MERCUSYS': 'TP-Link',
            'HUAWEI': 'Huawei', 'VELDS': 'Velds', 'FASTT10': 'Fastt10',
            'FIBRATECH': 'Fibratech', 'SANDISK': 'Sandisk',
            'SANDISK/WESTERN DIGITAL': 'Sandisk', 'ZYXEL': 'Zyxel',
            'XSIRIUS': 'XSirius', 'SINENG': 'Sineng', 'IP-COM': 'IpCom',
            'YOFC': 'Yofc', 'SKYLANE': 'Skylane', 'CO-NET': 'CoNet', 'NEP': 'Nep',
        }
        marca = marca_map.get(marca_norm) or detectar_marca(desc) or (marca_excel and str(marca_excel).strip()) or ''
        # Saldo atual: tenta primeiro pelo par (cod, nome_normalizado) — isola o caso
        # de mesmo cod / produtos diferentes por empresa. Se o cod já aparece com
        # DISCRIMINAÇÃO por nome no estoque, um nome que não bateu é 0 (não vaza
        # o total pra outro produto). Só cai pro cod puro se o cod não tem nome
        # diferenciado (dedup simples por CNPJ).
        chave_cn = (cod_int, _norm_nome(desc))
        if chave_cn in estoque_por_cod_nome:
            saldo = estoque_por_cod_nome[chave_cn]
        elif cod_int in cods_com_discriminacao:
            saldo = 0  # cod tem outros produtos com nome — esse não bateu
        else:
            saldo = estoque_por_codigo.get(cod_int, 0) or estoque_por_codigo_str.get(str(cod_int), 0)
        catalogo.append({
            "codigo": cod_int,
            "nome": str(desc).strip(),
            "marca": marca,
            "referencia": str(ref).strip() if ref and str(ref).strip() != '-' else '',
            "empresa_principal": EMPRESA_LABEL.get(empresa, f'Empresa {empresa}'),
            "saldo_atual": int(saldo) if isinstance(saldo, (int, float)) else 0,
            "preco_medio": 0.0,  # não disponível neste arquivo
        })
    catalogo.sort(key=lambda x: (x["nome"] or "").upper())
    print(f"  {len(catalogo)} produtos no catálogo (Sênior, fora do PSI; deduplicado por cod)")
else:
    # Fallback antigo: extrai do ESTOQUE SENIOR (só produtos com estoque > 0)
    print(f"  [AVISO] produtos_seniorx*.xlsx não encontrado — usando ESTOQUE SENIOR direto")
    for row in ws_estoque.iter_rows(min_row=2, values_only=True):
        cod = row[4]
        if cod is None or cod in codigos_psi:
            continue
        nome = row[5]
        filial = row[3]
        preco = row[13]
        qtd = row[9]
        if not nome:
            continue
        nome_norm = str(nome).strip()
        entry = next((c for c in catalogo if c["codigo"] == cod and c["nome"] == nome_norm), None)
        if entry is None:
            entry = {
                "codigo": cod, "nome": nome_norm,
                "marca": detectar_marca(nome_norm),
                "filial": str(filial).strip() if filial else "",
                "saldo_atual": 0,
                "preco_medio": float(preco) if isinstance(preco, (int, float)) else 0.0,
            }
            catalogo.append(entry)
        if isinstance(qtd, (int, float)):
            entry["saldo_atual"] += int(qtd)
    catalogo.sort(key=lambda x: (x["nome"] or "").upper())
    print(f"  {len(catalogo)} produtos no catálogo (fallback Sênior)")

# ---------------------------------------------------------------------------
# 7.5. Promove produto do catálogo → Curva A "virtual" quando overrides
#      tem fcst_mensal preenchido. Resolve o caso "novo produto no Sênior
#      que precisa virar PSI sem mexer no Excel".
#
# Promovido tem entry em `produtos[]` igual aos PSI normais:
#   - PSI semanal sintética (forecast = 0, propagado no cliente via
#     propagarFcstMensalParaPSI usando o próprio fcst_mensal)
#   - sell_in vindo do kanban
#   - saldo cascade a partir do estoque atual
#   - flag `_origem: 'catalogo_promovido'` pra UI distinguir
# ---------------------------------------------------------------------------
PRODUTO_OVERRIDES_PATH = os.path.join(os.path.dirname(__file__), "produto_overrides.json")
_overrides_data = {}
try:
    with open(PRODUTO_OVERRIDES_PATH, encoding="utf-8") as _f:
        _overrides_data = json.load(_f)
except FileNotFoundError:
    pass

# Semanas-padrão: usa o range completo do PSI Excel (W1-W53).
# Pega de um produto PSI existente pra garantir consistência.
_semanas_padrao = []
for _p in produtos_final:
    if _p.get("psi", {}).get("semanas"):
        _semanas_padrao = list(_p["psi"]["semanas"])
        break
if not _semanas_padrao:
    _semanas_padrao = list(range(1, 54))  # fallback W1-W53

_codigos_em_produtos = {p["codigo"] for p in produtos_final if p.get("codigo") is not None}
_catalogo_por_codigo = {c["codigo"]: c for c in catalogo if c.get("codigo") is not None}

_promovidos = 0
for _cod_str, _ov in _overrides_data.items():
    _fcst = _ov.get("fcst_mensal") or {}
    if not any(v for v in _fcst.values() if v):
        continue  # override sem forecast efetivo
    try: _cod = int(_cod_str)
    except (ValueError, TypeError): continue
    if _cod in _codigos_em_produtos: continue  # já é PSI normal — override aplica no cliente
    _cat = _catalogo_por_codigo.get(_cod)
    if _cat is None: continue  # nem está no catálogo — ignora

    # Saldo atual: pega do estoque consolidado (idem produtos PSI).
    _saldo = estoque_por_codigo.get(_cod, 0) or estoque_por_codigo.get(str(_cod), 0)
    # Fallback via MIGRACAO_CODIGO: produto recadastrado pode ter o saldo físico
    # cadastrado sob o código ANTIGO no Sênior (e via match por nome no PSI Excel).
    # Pega o saldo_atual já calculado do produto PSI antigo (que tem essa lógica
    # completa: estoque_por_codigo OR estoque_por_nome).
    if not _saldo and _cod in _MIGRACAO_CODIGO_REV:
        _cod_antigo = _MIGRACAO_CODIGO_REV[_cod]
        for _p_psi in produtos_final:
            if _p_psi.get("codigo") == _cod_antigo:
                _saldo = _p_psi.get("saldo_atual", 0)
                break

    # Sell_in semanal do kanban (match por codigo > nome)
    _sin_k = kanban_sin_por_codigo.get(_cod, {}) or \
             kanban_sin_por_nome.get((_cat.get("nome") or "").strip().upper(), {})
    # Sell_in EFETIVO (só chegadas físicas — fase estoque/recebido)
    _sin_k_efet = kanban_sin_efetivo_por_codigo.get(_cod, {}) or \
                  kanban_sin_efetivo_por_nome.get((_cat.get("nome") or "").strip().upper(), {})
    _ent_k = kanban_entradas_por_codigo.get(_cod, []) or \
             kanban_entradas_por_nome.get((_cat.get("nome") or "").strip().upper(), [])

    _forecast_arr = [0] * len(_semanas_padrao)  # cliente propaga via fcst_mensal
    _sellin_arr   = [int(_sin_k.get(w, 0)) for w in _semanas_padrao]
    _realizado_arr = [0] * len(_semanas_padrao)  # produto novo, sem histórico
    # BUG FIX (20/07/2026): aplica REALIZADO_MANUAL AQUI (antes do cascade) — o
    # post-processing rodava depois do cascade, entao o fix da W-1 nao acionava.
    if _cod in REALIZADO_MANUAL:
        for _sm, _qm in REALIZADO_MANUAL[_cod].items():
            if _sm in _semanas_padrao:
                _realizado_arr[_semanas_padrao.index(_sm)] = int(_qm)

    # Aplica saldo histórico (semanas fechadas congeladas em runs anteriores).
    # Sem isso, se um produto promovido teve saldo em W26 num run anterior mas
    # a SEMANA_ATUAL avançou pra W28, a W26 volta pra 0 (bug apontado 07/07).
    _hist_prod = HISTORICO_SALDO.get(str(_cod), {})

    # Cascade do saldo:
    # 1. Passado: usa histórico congelado quando existir; senão 0
    # 2. SEMANA_ESTOQUE em diante: saldo real + cascade forward
    _saldo_arr = [0] * len(_semanas_padrao)
    # Passado (semanas < SEMANA_ESTOQUE): usa histórico
    for _i, _w in enumerate(_semanas_padrao):
        if _w < SEMANA_ESTOQUE and str(_w) in _hist_prod:
            _saldo_arr[_i] = max(0, int(_hist_prod[str(_w)]))
    # SEMANA_ESTOQUE em diante: saldo real + cascade forward
    if SEMANA_ESTOQUE in _semanas_padrao:
        _idx_e = _semanas_padrao.index(SEMANA_ESTOQUE)
        if _idx_e > 0:
            # BUG FIX (20/07/2026): mesma logica do fix da 13/07 aplicada aos promovidos.
            # Se W anterior é FECHADA com realizado > 0, saldo final dela = conta
            # (nao a foto). Antes o cascade dos promovidos sempre usava a foto.
            _w_ant = _semanas_padrao[_idx_e - 1]
            _rl_ant = _realizado_arr[_idx_e - 1] if _idx_e - 1 < len(_realizado_arr) else 0
            if _w_ant < SEMANA_ATUAL and _rl_ant > 0 and _idx_e >= 2:
                _saldo_ant_ant = _saldo_arr[_idx_e - 2]
                # Usa sell_in EFETIVO — só chegadas realmente recebidas (fase estoque/recebido).
                # Se o pedido da W-1 ainda ta em produção/trânsito/numerário, a foto do
                # Sênior NÃO reflete essa chegada, entao somar sell_in inflaria o saldo.
                _si_ant = int(_sin_k_efet.get(_w_ant, 0))
                _saldo_arr[_idx_e - 1] = max(0, round(_saldo_ant_ant + _si_ant - _rl_ant))
            else:
                _saldo_arr[_idx_e - 1] = max(0, int(_saldo))
        # Regra de consumo (igual PSI Excel): semana fechada com realizado > 0 usa realizado.
        def _consumo_prom(i):
            _w = _semanas_padrao[i]
            _rl = _realizado_arr[i] if i < len(_realizado_arr) else 0
            _fc = _forecast_arr[i] if i < len(_forecast_arr) else 0
            return _rl if (_w <= SEMANA_ATUAL and _rl > 0) else _fc
        _saldo_arr[_idx_e] = max(0, int(_saldo) - _consumo_prom(_idx_e) + _sellin_arr[_idx_e])
        for _i in range(_idx_e + 1, len(_semanas_padrao)):
            _saldo_arr[_i] = max(0, _saldo_arr[_i-1] + _sellin_arr[_i] - _consumo_prom(_i))

    # Congela no histórico o saldo final de todas as semanas <= SEMANA_ESTOQUE
    # pra próximas rodadas preservarem quando SEMANA_ATUAL avançar.
    _novo_hist = dict(_hist_prod)
    for _i, _w in enumerate(_semanas_padrao):
        if _w <= SEMANA_ESTOQUE:
            _novo_hist[str(_w)] = max(0, int(_saldo_arr[_i]))
    HISTORICO_SALDO[str(_cod)] = _novo_hist

    # Entradas futuras
    _ent_futuras = sorted([e for e in _ent_k if e.get("semana", 0) >= SEMANA_ATUAL],
                          key=lambda e: e["semana"])
    _prox_chegada = _ent_futuras[0] if _ent_futuras else None
    # in_transit: mesma regra do bloco principal (fases embarque em diante)
    _FASES_EM_TRANSITO_2 = {"embarque", "porto", "numerario", "transito"}
    _in_transit = sum(e["quantidade"] for e in _ent_futuras
                      if (e.get("fase") or "").lower() in _FASES_EM_TRANSITO_2)

    # Forecast semanal médio: 3 meses iniciais ÷ semanas
    MESES_ORDEM = [
        'MAI/26','JUN/26','JUL/26','AGO/26','SET/26','OUT/26',
        'NOV/26','DEZ/26','JAN/27','FEV/27','MAR/27','ABR/27','MAI/27',
    ]
    _meses_com_fc = [m for m in MESES_ORDEM if _fcst.get(m, 0) > 0][:3]
    if _meses_com_fc:
        _total_fc  = sum(_fcst.get(m, 0) for m in _meses_com_fc)
        _total_sem = sum(semanas_do_mes(m) for m in _meses_com_fc)
        _fc_sem = round(_total_fc / _total_sem) if _total_sem > 0 else 0
    else:
        _fc_sem = 0

    produtos_final.append({
        "nome":            _cat.get("nome", ""),
        "marca":           _ov.get("marca") or _cat.get("marca", ""),
        "codigo":          _cod,
        "familia":         _cat.get("familia", "") or "",
        "tipo":            _ov.get("tipo") or "Nacional",
        "pv":              float(_ov.get("pv", 0) or 0),
        "pc":              float(_ov.get("pc", 0) or 0),
        "saldo_atual":     int(_saldo) if isinstance(_saldo, (int, float)) else 0,
        "fc_sem":          _fc_sem,
        "in_transit":      _in_transit,
        "prox_chegada":    _prox_chegada,
        "entradas":        _ent_futuras,
        "fcst_mensal":     _fcst,
        "venda_mensal":    {},
        "realizado_split": {},
        "psi": {
            "semanas":   _semanas_padrao,
            "forecast":  _forecast_arr,
            "sell_in":   _sellin_arr,
            "realizado": _realizado_arr,
            "saldo":     _saldo_arr,
        },
        "_origem": "catalogo_promovido",
    })
    _promovidos += 1

if _promovidos:
    print(f"  Promovidos do catalogo (override.fcst_mensal): {_promovidos} produtos")

# Remove do catálogo os que viraram Curva A (evita duplicidade na aba Catálogo)
_codigos_promovidos = {p["codigo"] for p in produtos_final if p.get("_origem") == "catalogo_promovido"}
catalogo = [c for c in catalogo if c.get("codigo") not in _codigos_promovidos]

# Post-processing: aplica REALIZADO_MANUAL em TODOS os produtos_final (inclusive
# os promovidos do catálogo, que o loop principal não pegou porque não vieram
# do PSI Excel). Sobrescreve psi.realizado (total) e realizado_split (por mês).
for _p in produtos_final:
    _c = _p.get("codigo")
    if _c not in REALIZADO_MANUAL: continue
    _psi = _p.get("psi")
    if _psi and isinstance(_psi.get("semanas"), list) and isinstance(_psi.get("realizado"), list):
        for _sem_alvo, _qtd in REALIZADO_MANUAL[_c].items():
            if _sem_alvo in _psi["semanas"]:
                _psi["realizado"][_psi["semanas"].index(_sem_alvo)] = int(_qtd)
    # Popula realizado_split: primeiro deriva do REALIZADO_MANUAL (total no mês-dono
    # da semana), depois SOBRESCREVE com REALIZADO_SPLIT_MANUAL se existir (pra
    # semanas-fronteira dividirem certo entre 2 meses).
    _rs = dict(_p.get("realizado_split") or {})
    # Limpa entradas antigas de semanas que estão em REALIZADO_MANUAL — evita
    # split "fantasma" da rodada anterior (ex: 27_JUL/26=152 aparecendo em vez de 122)
    _semanas_manual = set(REALIZADO_MANUAL[_c].keys())
    _rs = {k: v for k, v in _rs.items() if not any(k.startswith(f"{s}_") for s in _semanas_manual)}
    for _sem_alvo, _qtd in REALIZADO_MANUAL[_c].items():
        _mes_alvo = WEEK_TO_MONTH.get(_sem_alvo)
        if _mes_alvo is None:
            for _w, _m in sorted(WEEK_TO_MONTH.items()):
                if _w <= _sem_alvo: _mes_alvo = _m
        if _mes_alvo:
            _rs[f"{_sem_alvo}_{_mes_alvo}"] = int(_qtd)
    # Aplica split explícito por cima (sobrescreve o total colocado no mês-dono)
    if _c in REALIZADO_SPLIT_MANUAL:
        for _chave, _qtd in REALIZADO_SPLIT_MANUAL[_c].items():
            _rs[_chave] = int(_qtd)
    _p["realizado_split"] = _rs

# ---------------------------------------------------------------------------
# 8. Gerar dados.json
# ---------------------------------------------------------------------------
try:
    _linhas_est = linhas_estoque
except NameError:
    _linhas_est = []

# Data do arquivo posicao_estoque_todas — usada pra mostrar "Atualizado em DD/MM/YYYY"
_estoque_data_iso = None
_estoque_fonte = None
if POSICAO_TODAS_PATH:
    import re as _re
    _m = _re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(POSICAO_TODAS_PATH))
    if _m: _estoque_data_iso = _m.group(1)
    _estoque_fonte = os.path.basename(POSICAO_TODAS_PATH)

output = {
    "gerado_em":    datetime.datetime.now().isoformat(timespec="seconds"),
    "semana_atual": SEMANA_ATUAL,
    "semana_estoque": SEMANA_ESTOQUE,
    "week_to_month": {str(k): v for k, v in WEEK_TO_MONTH.items()},
    "produtos":     produtos_final,
    "catalogo":     catalogo,
    "vendas_b2b":   vendas_b2b_linhas,
    "estoque":      _linhas_est,
    "estoque_atualizado_em": _estoque_data_iso,
    "estoque_fonte": _estoque_fonte,
}

with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
    json.dump(output, f, ensure_ascii=False, indent=2, default=str)

# Salva histórico de saldo (semanas fechadas congeladas)
with open(HISTORICO_FILE, "w", encoding="utf-8") as f:
    json.dump(HISTORICO_SALDO, f, ensure_ascii=False, indent=2)

print(f"\nArquivo gerado: {OUTPUT_FILE}")
print(f"Histórico de saldo: {HISTORICO_FILE} ({sum(len(v) for v in HISTORICO_SALDO.values())} pontos congelados)")
print(f"Total produtos: {len(produtos_final)}")
